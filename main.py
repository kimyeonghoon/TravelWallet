"""
일본 여행 경비 추적 애플리케이션 메인 서버

이 애플리케이션은 일본 여행 중 발생하는 경비를 추적하고 관리하는 웹 애플리케이션입니다.
주요 기능:
- 지출 내역 관리 (CRUD)
- 교통카드 잔액 관리
- 엔화 지갑 관리
- 실시간 환율 연동 (한국수출입은행 API)
- 텔레그램 봇을 통한 인증 시스템
- 통계 및 데이터 내보내기

기술 스택: FastAPI, SQLite, Bootstrap 5, jQuery
"""

# FastAPI 및 관련 라이브러리 임포트
from fastapi import FastAPI, Request, Depends, HTTPException, Cookie
from fastapi.responses import HTMLResponse, JSONResponse
from fastapi.staticfiles import StaticFiles  # 정적 파일 서빙
from fastapi.templating import Jinja2Templates  # HTML 템플릿 렌더링
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials  # JWT 인증
from sqlalchemy.orm import Session  # 데이터베이스 세션 관리
from pydantic import BaseModel  # 데이터 검증 모델
from typing import List, Optional  # 타입 힌팅
from contextlib import asynccontextmanager  # Lifespan events
import os

# 자체 모듈 임포트
from models import create_tables, get_db, User  # 데이터베이스 모델
from database import TripService, ExpenseService, TransportCardService, WalletService, TransportationService  # 데이터베이스 서비스
from auth import AuthService  # 인증 서비스
from exchange_service import exchange_service  # 환율 서비스
from models import engine, Expense  # 마이그레이션을 위한 모델 임포트

def migrate_existing_expenses_to_default_trip(db: Session):
    """
    기존 trip_id가 None인 지출들을 기본 여행으로 마이그레이션
    """
    try:
        # 기본 여행 생성 또는 가져오기
        default_trip = TripService.create_default_trip_if_not_exists(db)

        # trip_id가 None인 지출들 조회
        expenses_without_trip = db.query(Expense).filter(Expense.trip_id.is_(None)).all()

        if expenses_without_trip:
            print(f"기존 {len(expenses_without_trip)}건의 지출을 기본 여행({default_trip.name})으로 마이그레이션 중...")

            # 모든 지출의 trip_id를 기본 여행 ID로 업데이트
            for expense in expenses_without_trip:
                expense.trip_id = default_trip.id

            db.commit()
            print(f"마이그레이션 완료: {len(expenses_without_trip)}건의 지출이 '기본 여행'으로 이동되었습니다")
        else:
            print("마이그레이션할 지출 데이터가 없습니다")

    except Exception as e:
        print(f"마이그레이션 오류: {str(e)}")
        db.rollback()

@asynccontextmanager
async def lifespan(app: FastAPI):
    # Startup
    create_tables()
    print("데이터베이스 테이블이 성공적으로 생성되었습니다")

    # 기존 지출 데이터를 기본 여행으로 마이그레이션
    from sqlalchemy.orm import sessionmaker
    SessionLocal = sessionmaker(autocommit=False, autoflush=False, bind=engine)
    db = SessionLocal()
    try:
        migrate_existing_expenses_to_default_trip(db)
    finally:
        db.close()

    yield
    # Shutdown (if needed)

# FastAPI 애플리케이션 인스턴스 생성
app = FastAPI(
    title="Japan Travel Expense Tracker",
    description="일본 여행 경비 추적 시스템",
    lifespan=lifespan
)

# CORS 미들웨어 추가 (nginx 프록시 호환성을 위함)
from fastapi.middleware.cors import CORSMiddleware
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],  # 프로덕션에서는 특정 도메인으로 제한 필요
    allow_credentials=True,  # 쿠키 및 인증 정보 허용
    allow_methods=["*"],  # 모든 HTTP 메서드 허용
    allow_headers=["*"],  # 모든 헤더 허용
)

# 정적 파일 및 템플릿 설정
static_dir = os.path.join(os.path.dirname(__file__), "static")  # 정적 파일 디렉토리 경로
app.mount("/static", StaticFiles(directory=static_dir, html=True), name="static")  # CSS, JS, 이미지 파일 서빙
templates = Jinja2Templates(directory="templates")  # HTML 템플릿 디렉토리 설정

# Lifespan 이벤트로 대체됨 - 위의 lifespan 함수 참조

# ==================== API 요청/응답 모델 정의 (Pydantic) ====================

# ==================== 여행 관련 모델 ====================

class TripCreate(BaseModel):
    """여행 생성 요청 모델"""
    name: str  # 여행 이름
    destination: str  # 여행지
    start_date: str  # 시작일 (YYYY-MM-DD 형식)
    end_date: str  # 종료일 (YYYY-MM-DD 형식)
    description: str = ""  # 여행 설명 (선택사항)

class TripUpdate(BaseModel):
    """여행 수정 요청 모델 (모든 필드 선택사항)"""
    name: Optional[str] = None  # 수정할 여행 이름
    destination: Optional[str] = None  # 수정할 여행지
    start_date: Optional[str] = None  # 수정할 시작일
    end_date: Optional[str] = None  # 수정할 종료일
    description: Optional[str] = None  # 수정할 여행 설명

class TripResponse(BaseModel):
    """여행 조회 응답 모델"""
    id: int  # 여행 고유 ID
    name: str  # 여행 이름
    destination: str  # 여행지
    start_date: str  # 시작일
    end_date: str  # 종료일
    description: str  # 여행 설명
    is_default: bool  # 기본 여행 여부
    created_at: str  # 생성 시간
    updated_at: str  # 수정 시간

# ==================== 지출 관련 모델 ====================

class ExpenseCreate(BaseModel):
    """지출 생성 요청 모델"""
    amount: float  # 지출 금액
    category: str  # 지출 카테고리 (식비, 교통비, 숙박비, 입장료, 기타)
    description: str = ""  # 지출 설명 (선택사항)
    payment_method: str = "현금"  # 결제 수단 (현금, 체크카드, 신용카드, 교통카드)
    wallet_id: Optional[int] = None  # 지갑 ID (현금 결제 시 선택사항)
    trip_id: Optional[int] = None  # 여행 ID (여행별 지출 분류)

class ExpenseResponse(BaseModel):
    """지출 조회 응답 모델"""
    id: int  # 지출 고유 ID
    amount: float  # 지출 금액
    category: str  # 지출 카테고리
    description: str  # 지출 설명
    date: str  # 지출 날짜 (YYYY-MM-DD 형식)
    payment_method: str  # 결제 수단
    wallet_id: Optional[int] = None  # 지갑 ID
    wallet_name: Optional[str] = None  # 지갑 이름
    timestamp: str  # 등록 시간 (ISO 형식)

class ExpenseUpdate(BaseModel):
    """지출 수정 요청 모델 (모든 필드 선택사항)"""
    amount: Optional[float] = None  # 수정할 금액
    category: Optional[str] = None  # 수정할 카테고리
    description: Optional[str] = None  # 수정할 설명
    date: Optional[str] = None  # 수정할 날짜
    time: Optional[str] = None  # 수정할 시간
    payment_method: Optional[str] = None  # 수정할 결제 수단
    wallet_id: Optional[int] = None  # 수정할 지갑 ID

class SummaryResponse(BaseModel):
    """지출 요약 정보 응답 모델"""
    total_expense: float  # 총 지출 금액
    today_expense: float  # 오늘 지출 금액

# ==================== 교통카드 관련 모델 ====================

class TransportCardCreate(BaseModel):
    """교통카드 생성 요청 모델"""
    name: str  # 교통카드 이름 (예: 스이카, 파스모 등)
    balance: float = 0.0  # 초기 잔액 (엔화)

class TransportCardUpdate(BaseModel):
    name: Optional[str] = None
    balance: Optional[float] = None

class TransportCardResponse(BaseModel):
    id: int
    name: str
    balance: float
    created_at: str
    updated_at: str

# Wallet models
class WalletCreate(BaseModel):
    name: str
    balance: float = 0.0

class WalletUpdate(BaseModel):
    name: Optional[str] = None
    balance: Optional[float] = None

class WalletResponse(BaseModel):
    id: int
    name: str
    balance: float
    created_at: str
    updated_at: str

# ==================== 교통수단 관련 모델 ====================

class TransportationCreate(BaseModel):
    """교통수단 기록 생성 요청 모델"""
    category: str  # 교통수단 카테고리 (JR, 전철, 버스, 배, 기타)
    company: str = ""  # 이용회사 (서일본, 히로덴 등)
    departure_time: str  # 출발시간 (HH:MM)
    arrival_time: str  # 도착시간 (HH:MM)
    memo: str = ""  # 메모 (출발지-도착지, 노선 등)

class TransportationUpdate(BaseModel):
    """교통수단 기록 수정 요청 모델 (모든 필드 선택사항)"""
    category: Optional[str] = None
    company: Optional[str] = None
    departure_time: Optional[str] = None
    arrival_time: Optional[str] = None
    memo: Optional[str] = None
    date: Optional[str] = None

class TransportationResponse(BaseModel):
    """교통수단 기록 조회 응답 모델"""
    id: int
    user_id: Optional[int] = None
    category: str
    company: str
    departure_time: str
    arrival_time: str
    memo: str
    date: str
    timestamp: str

# Authentication models
class LoginRequest(BaseModel):
    email: str

class LoginCodeRequest(BaseModel):
    code: str

class LoginResponse(BaseModel):
    message: str

class UserResponse(BaseModel):
    id: int
    telegram_chat_id: str
    email: Optional[str] = None
    is_active: bool

# Security
security = HTTPBearer(auto_error=False)

def get_client_ip(request: Request) -> str:
    """Get client IP address from request."""
    # Check for forwarded header (when behind proxy)
    forwarded_for = request.headers.get("X-Forwarded-For")
    if forwarded_for:
        # Take the first IP in the list
        return forwarded_for.split(',')[0].strip()
    
    # Check for real IP header
    real_ip = request.headers.get("X-Real-IP")
    if real_ip:
        return real_ip.strip()
    
    # Fall back to direct connection IP
    if request.client:
        return request.client.host
    
    return "127.0.0.1"  # Fallback

async def get_current_user(
    request: Request,
    credentials: HTTPAuthorizationCredentials = Depends(security),
    session_token: str = Cookie(None),
    db: Session = Depends(get_db)
) -> Optional[User]:
    """Get current authenticated user from JWT token."""
    token = None
    
    if credentials:
        token = credentials.credentials
    elif session_token:
        token = session_token
    
    if not token:
        return None
    
    payload = AuthService.verify_token(token)
    if not payload:
        return None
    
    user_id = payload.get("user_id")
    if not user_id:
        return None
    
    user = db.query(User).filter(User.id == user_id).first()
    return user if user and user.is_active else None

async def require_auth(current_user: User = Depends(get_current_user)) -> User:
    """Require authentication for protected routes."""
    if not current_user:
        raise HTTPException(status_code=401, detail="Authentication required")
    return current_user

@app.get("/", response_class=HTMLResponse)
async def read_root(
    request: Request, 
    current_user: Optional[User] = Depends(get_current_user)
):
    """Main page accessible to all users, with login features for authenticated users."""
    return templates.TemplateResponse("index.html", {
        "request": request,
        "user": current_user
    })

@app.get("/statistics", response_class=HTMLResponse)
async def statistics_page(
    request: Request,
    current_user: Optional[User] = Depends(get_current_user)
):
    """Statistics dashboard page accessible to all users."""
    return templates.TemplateResponse("statistics.html", {
        "request": request,
        "user": current_user
    })

@app.get("/transport-cards", response_class=HTMLResponse)
async def transport_cards_page(
    request: Request,
    current_user: Optional[User] = Depends(get_current_user)
):
    """Transport cards page accessible to all users."""
    return templates.TemplateResponse("transport-cards.html", {
        "request": request,
        "user": current_user
    })

@app.get("/transportation", response_class=HTMLResponse)
async def transportation_page(
    request: Request,
    current_user: Optional[User] = Depends(get_current_user)
):
    """Transportation records page accessible to all users."""
    return templates.TemplateResponse("transportation.html", {
        "request": request,
        "user": current_user
    })

@app.get("/api/health")
async def health_check():
    return {"status": "ok", "message": "Japan Travel Expense API is running"}

# Add explicit static file handling for nginx compatibility
from fastapi.responses import FileResponse
import mimetypes

@app.get("/static/{file_path:path}")
async def serve_static_files(file_path: str):
    """Serve static files with proper MIME types for nginx compatibility."""
    full_path = os.path.join(static_dir, file_path)
    
    if not os.path.exists(full_path):
        raise HTTPException(status_code=404, detail="File not found")
    
    # Get MIME type
    mime_type, _ = mimetypes.guess_type(full_path)
    if mime_type is None:
        mime_type = "application/octet-stream"
    
    return FileResponse(
        path=full_path,
        media_type=mime_type,
        headers={
            "Cache-Control": "public, max-age=31536000",  # 1 year cache
            "Access-Control-Allow-Origin": "*"
        }
    )


# Authentication endpoints
@app.post("/api/auth/login", response_model=LoginResponse)
async def request_login_code(
    login_data: LoginRequest, 
    request: Request,
    db: Session = Depends(get_db)
):
    """Request login code via email verification."""
    try:
        # Get client IP
        client_ip = get_client_ip(request)
        
        # Verify email and send code
        success, message = AuthService.verify_email_and_send_code(
            db, login_data.email, client_ip
        )
        
        if not success:
            raise HTTPException(status_code=400, detail=message)
        
        return LoginResponse(message=message)
        
    except HTTPException:
        raise
    except Exception as e:
        print(f"Failed to send login code: {e}")
        raise HTTPException(status_code=500, detail="로그인 코드 전송에 실패했습니다.")

@app.post("/api/auth/verify")
async def verify_login_code(code_data: LoginCodeRequest, db: Session = Depends(get_db)):
    """Verify login code and create session."""
    user = AuthService.validate_login_code(db, code_data.code)
    
    if not user:
        raise HTTPException(status_code=400, detail="Invalid or expired code")
    
    # Create JWT access token
    access_token = AuthService.create_access_token({"user_id": user.id})
    
    # Return success with token
    response = JSONResponse({"message": "Login successful", "user_id": user.id})
    response.set_cookie(
        key="session_token",
        value=access_token,
        httponly=True,
        max_age=900,  # 15 minutes
        secure=False  # Set to True in production with HTTPS
    )
    
    return response


# Removed login page route - login is now handled via modal in main page

@app.post("/api/auth/logout")
async def logout():
    """Logout user by clearing session cookie."""
    response = JSONResponse({"message": "Successfully logged out"})
    response.delete_cookie("session_token")
    return response

@app.get("/api/auth/me", response_model=UserResponse)
async def get_current_user_info(current_user: User = Depends(get_current_user)):
    """Get current user information."""
    if not current_user:
        raise HTTPException(status_code=401, detail="Not authenticated")
    
    return UserResponse(
        id=current_user.id,
        telegram_chat_id=current_user.telegram_chat_id,
        email=current_user.email,
        is_active=current_user.is_active
    )

# ==================== 여행 관리 API ====================

@app.post("/api/trips", response_model=TripResponse)
async def create_trip(trip: TripCreate, current_user: User = Depends(require_auth), db: Session = Depends(get_db)):
    """새로운 여행을 생성합니다."""
    try:
        new_trip = TripService.create_trip(
            db=db,
            name=trip.name,
            destination=trip.destination,
            start_date=trip.start_date,
            end_date=trip.end_date,
            description=trip.description
        )
        return TripResponse(**new_trip.to_dict())
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/trips", response_model=List[TripResponse])
async def get_trips(db: Session = Depends(get_db)):
    """모든 여행 목록을 조회합니다."""
    try:
        trips = TripService.get_all_trips(db)
        return [TripResponse(**trip.to_dict()) for trip in trips]
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/trips/{trip_id}", response_model=TripResponse)
async def get_trip(trip_id: int, db: Session = Depends(get_db)):
    """특정 여행 정보를 조회합니다."""
    try:
        trip = TripService.get_trip_by_id(db, trip_id)
        if not trip:
            raise HTTPException(status_code=404, detail="Trip not found")
        return TripResponse(**trip.to_dict())
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.put("/api/trips/{trip_id}", response_model=TripResponse)
async def update_trip(trip_id: int, trip_update: TripUpdate, current_user: User = Depends(require_auth), db: Session = Depends(get_db)):
    """여행 정보를 수정합니다."""
    try:
        updated_trip = TripService.update_trip(
            db=db,
            trip_id=trip_id,
            name=trip_update.name,
            destination=trip_update.destination,
            start_date=trip_update.start_date,
            end_date=trip_update.end_date,
            description=trip_update.description
        )
        if not updated_trip:
            raise HTTPException(status_code=404, detail="Trip not found")
        return TripResponse(**updated_trip.to_dict())
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.delete("/api/trips/{trip_id}")
async def delete_trip(trip_id: int, current_user: User = Depends(require_auth), db: Session = Depends(get_db)):
    """여행을 삭제합니다."""
    try:
        success = TripService.delete_trip(db, trip_id)
        if not success:
            raise HTTPException(status_code=404, detail="Trip not found or cannot delete default trip")
        return {"message": "Trip deleted successfully"}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/trips/{trip_id}/set-default")
async def set_default_trip(trip_id: int, current_user: User = Depends(require_auth), db: Session = Depends(get_db)):
    """기본 여행을 설정합니다."""
    try:
        success = TripService.set_default_trip(db, trip_id)
        if not success:
            raise HTTPException(status_code=404, detail="Trip not found")
        return {"message": "Default trip set successfully"}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

# ==================== 지출 관리 API ====================

@app.post("/api/expenses", response_model=ExpenseResponse)
async def create_expense(expense: ExpenseCreate, current_user: User = Depends(require_auth), db: Session = Depends(get_db)):
    """Create a new expense."""
    try:
        new_expense = ExpenseService.create_expense(
            db=db,
            user_id=current_user.id,
            amount=expense.amount,
            category=expense.category,
            description=expense.description,
            payment_method=expense.payment_method,
            wallet_id=expense.wallet_id,
            trip_id=expense.trip_id
        )
        return ExpenseResponse(**new_expense.to_dict())
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/expenses", response_model=List[ExpenseResponse])
async def get_expenses(
    category: Optional[str] = None,
    payment_method: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    sort_by: Optional[str] = None,
    sort_order: Optional[str] = "desc",
    search: Optional[str] = None,
    db: Session = Depends(get_db)
):
    """Get expenses with optional filters and sorting - public access for expense viewing."""
    if any([category, payment_method, date_from, date_to, sort_by, search]):
        # Use filtered query
        expenses = ExpenseService.get_filtered_expenses(
            db=db,
            category=category,
            payment_method=payment_method,
            date_from=date_from,
            date_to=date_to,
            sort_by=sort_by,
            sort_order=sort_order,
            search=search
        )
    else:
        # Use existing method for backward compatibility
        expenses = ExpenseService.get_all_expenses(db)

    return [ExpenseResponse(**expense.to_dict()) for expense in expenses]

@app.get("/api/expenses/by-date/{date}")
async def get_expenses_by_date(
    date: str,
    db: Session = Depends(get_db)
):
    """Get expenses for a specific date - public access for expense viewing."""
    try:
        # Use existing filtered method with date range
        expenses = ExpenseService.get_filtered_expenses(
            db=db,
            date_from=date,
            date_to=date,
            sort_by="created_at",
            sort_order="desc"
        )
        return [ExpenseResponse(**expense.to_dict()) for expense in expenses]
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.put("/api/expenses/{expense_id}", response_model=ExpenseResponse)
async def update_expense(expense_id: int, expense_update: ExpenseUpdate, current_user: User = Depends(require_auth), db: Session = Depends(get_db)):
    """Update an expense."""
    updated_expense = ExpenseService.update_user_expense(
        db=db,
        user_id=current_user.id,
        expense_id=expense_id,
        amount=expense_update.amount,
        category=expense_update.category,
        description=expense_update.description,
        expense_date=expense_update.date,
        expense_time=expense_update.time,
        payment_method=expense_update.payment_method,
        wallet_id=expense_update.wallet_id
    )
    if not updated_expense:
        raise HTTPException(status_code=404, detail="Expense not found")
    return ExpenseResponse(**updated_expense.to_dict())

@app.delete("/api/expenses/{expense_id}")
async def delete_expense(expense_id: int, current_user: User = Depends(require_auth), db: Session = Depends(get_db)):
    """Delete an expense."""
    success = ExpenseService.delete_user_expense(db, current_user.id, expense_id)
    if not success:
        raise HTTPException(status_code=404, detail="Expense not found")
    return {"message": "Expense deleted successfully"}

@app.get("/api/summary", response_model=SummaryResponse)
async def get_summary(db: Session = Depends(get_db)):
    """Get expense summary - public access for viewing totals."""
    total_expense = ExpenseService.get_total_expenses(db)
    today_expense = ExpenseService.get_today_expenses_total(db)
    
    return SummaryResponse(
        total_expense=total_expense,
        today_expense=today_expense
    )

@app.get("/api/statistics")
async def get_statistics(db: Session = Depends(get_db)):
    """Get comprehensive statistics for dashboard."""
    return ExpenseService.get_statistics(db)

# Transport Card endpoints
@app.post("/api/transport-cards", response_model=TransportCardResponse)
async def create_transport_card(
    card: TransportCardCreate,
    _: User = Depends(require_auth),
    db: Session = Depends(get_db)
):
    """Create a new transport card."""
    try:
        new_card = TransportCardService.create_card(db, card.name, card.balance)
        return TransportCardResponse(**new_card.to_dict())
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/transport-cards", response_model=List[TransportCardResponse])
async def get_transport_cards(db: Session = Depends(get_db)):
    """Get all transport cards - public access for viewing."""
    cards = TransportCardService.get_all_cards(db)
    return [TransportCardResponse(**card.to_dict()) for card in cards]

@app.put("/api/transport-cards/{card_id}", response_model=TransportCardResponse)
async def update_transport_card(
    card_id: int,
    card_update: TransportCardUpdate,
    _: User = Depends(require_auth),
    db: Session = Depends(get_db)
):
    """Update a transport card."""
    updated_card = TransportCardService.update_card(
        db, card_id, card_update.name, card_update.balance
    )
    if not updated_card:
        raise HTTPException(status_code=404, detail="Transport card not found")
    return TransportCardResponse(**updated_card.to_dict())

@app.delete("/api/transport-cards/{card_id}")
async def delete_transport_card(
    card_id: int,
    _: User = Depends(require_auth),
    db: Session = Depends(get_db)
):
    """Delete a transport card."""
    success = TransportCardService.delete_card(db, card_id)
    if not success:
        raise HTTPException(status_code=404, detail="Transport card not found")
    return {"message": "Transport card deleted successfully"}

@app.get("/api/transport-cards/summary")
async def get_transport_card_summary(db: Session = Depends(get_db)):
    """Get total balance of all transport cards."""
    total_balance = TransportCardService.get_total_balance(db)
    return {"total_balance": total_balance}

# Wallet endpoints
@app.get("/wallets", response_class=HTMLResponse)
async def wallets_page(
    request: Request,
    current_user: Optional[User] = Depends(get_current_user)
):
    """Wallets page accessible to all users."""
    return templates.TemplateResponse("wallets.html", {
        "request": request,
        "user": current_user
    })

@app.post("/api/wallets", response_model=WalletResponse)
async def create_wallet(
    wallet: WalletCreate,
    _: User = Depends(require_auth),
    db: Session = Depends(get_db)
):
    """Create a new wallet."""
    try:
        new_wallet = WalletService.create_wallet(db, wallet.name, wallet.balance)
        return WalletResponse(**new_wallet.to_dict())
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/wallets", response_model=List[WalletResponse])
async def get_wallets(db: Session = Depends(get_db)):
    """Get all wallets - public access for viewing."""
    wallets = WalletService.get_all_wallets(db)
    return [WalletResponse(**wallet.to_dict()) for wallet in wallets]

@app.put("/api/wallets/{wallet_id}", response_model=WalletResponse)
async def update_wallet(
    wallet_id: int,
    wallet_update: WalletUpdate,
    _: User = Depends(require_auth),
    db: Session = Depends(get_db)
):
    """Update a wallet."""
    updated_wallet = WalletService.update_wallet(
        db, wallet_id, wallet_update.name, wallet_update.balance
    )
    if not updated_wallet:
        raise HTTPException(status_code=404, detail="Wallet not found")
    return WalletResponse(**updated_wallet.to_dict())

@app.delete("/api/wallets/{wallet_id}")
async def delete_wallet(
    wallet_id: int,
    _: User = Depends(require_auth),
    db: Session = Depends(get_db)
):
    """Delete a wallet."""
    success = WalletService.delete_wallet(db, wallet_id)
    if not success:
        raise HTTPException(status_code=404, detail="Wallet not found")
    return {"message": "Wallet deleted successfully"}

@app.get("/api/wallets/summary")
async def get_wallet_summary(db: Session = Depends(get_db)):
    """Get total balance of all wallets."""
    total_balance = WalletService.get_total_balance(db)
    return {"total_balance": total_balance}

# ==================== 교통수단 API 엔드포인트 ====================

@app.post("/api/transportation", response_model=TransportationResponse)
async def create_transportation(
    transportation: TransportationCreate,
    current_user: User = Depends(require_auth),
    db: Session = Depends(get_db)
):
    """Create a new transportation record."""
    try:
        new_transportation = TransportationService.create_transportation(
            db=db,
            user_id=current_user.id,
            category=transportation.category,
            company=transportation.company,
            departure_time=transportation.departure_time,
            arrival_time=transportation.arrival_time,
            memo=transportation.memo
        )
        return TransportationResponse(**new_transportation.to_dict())
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/transportation", response_model=List[TransportationResponse])
async def get_transportation_records(
    category: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    sort_by: Optional[str] = None,
    sort_order: Optional[str] = "desc",
    db: Session = Depends(get_db)
):
    """Get transportation records with optional filters and sorting - public access for viewing."""
    if any([category, date_from, date_to, sort_by]):
        # Use filtered query
        transportations = TransportationService.get_filtered_transportations(
            db=db,
            category=category,
            date_from=date_from,
            date_to=date_to,
            sort_by=sort_by,
            sort_order=sort_order
        )
    else:
        # Use existing method for backward compatibility
        transportations = TransportationService.get_all_transportations(db)

    return [TransportationResponse(**transportation.to_dict()) for transportation in transportations]

@app.put("/api/transportation/{transportation_id}", response_model=TransportationResponse)
async def update_transportation(
    transportation_id: int,
    transportation_update: TransportationUpdate,
    current_user: User = Depends(require_auth),
    db: Session = Depends(get_db)
):
    """Update a transportation record."""
    updated_transportation = TransportationService.update_user_transportation(
        db=db,
        user_id=current_user.id,
        transportation_id=transportation_id,
        category=transportation_update.category,
        company=transportation_update.company,
        departure_time=transportation_update.departure_time,
        arrival_time=transportation_update.arrival_time,
        memo=transportation_update.memo,
        transportation_date=transportation_update.date
    )
    if not updated_transportation:
        raise HTTPException(status_code=404, detail="Transportation record not found")
    return TransportationResponse(**updated_transportation.to_dict())

@app.delete("/api/transportation/{transportation_id}")
async def delete_transportation(
    transportation_id: int,
    current_user: User = Depends(require_auth),
    db: Session = Depends(get_db)
):
    """Delete a transportation record."""
    success = TransportationService.delete_user_transportation(db, current_user.id, transportation_id)
    if not success:
        raise HTTPException(status_code=404, detail="Transportation record not found")
    return {"message": "Transportation record deleted successfully"}

# Exchange Rate endpoints
@app.get("/api/exchange-rate")
async def get_exchange_rate():
    """Get current JPY to KRW exchange rate."""
    try:
        rate_info = exchange_service.get_rate_info()
        return rate_info
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to fetch exchange rate: {str(e)}")

@app.post("/api/convert/jpy-to-krw")
async def convert_jpy_to_krw(amount: dict):
    """Convert JPY amount to KRW."""
    try:
        jpy_amount = amount.get("amount", 0)
        krw_amount = exchange_service.convert_jpy_to_krw(jpy_amount)
        return {
            "jpy_amount": jpy_amount,
            "krw_amount": krw_amount,
            "exchange_rate": exchange_service.get_jpy_to_krw_rate()
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Conversion failed: {str(e)}")

@app.post("/api/convert/krw-to-jpy")
async def convert_krw_to_jpy(amount: dict):
    """Convert KRW amount to JPY."""
    try:
        krw_amount = amount.get("amount", 0)
        jpy_amount = exchange_service.convert_krw_to_jpy(krw_amount)
        return {
            "krw_amount": krw_amount,
            "jpy_amount": jpy_amount,
            "exchange_rate": exchange_service.get_jpy_to_krw_rate()
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Conversion failed: {str(e)}")

# Data Export endpoints
@app.get("/api/export/csv")
async def export_expenses_csv(
    category: Optional[str] = None,
    payment_method: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    db: Session = Depends(get_db)
):
    """Export expenses data as CSV file."""
    try:
        import csv
        import io
        from datetime import datetime
        
        # Get filtered expenses
        expenses = ExpenseService.get_filtered_expenses(
            db=db,
            category=category,
            payment_method=payment_method,
            date_from=date_from,
            date_to=date_to,
            sort_by="date",
            sort_order="desc"
        )
        
        # Get current exchange rate for conversion
        try:
            exchange_rate = exchange_service.get_jpy_to_krw_rate()
        except:
            exchange_rate = 9.5  # Fallback rate
        
        # Create CSV content
        output = io.StringIO()
        writer = csv.writer(output)
        
        # Write header with filter info
        if any([category, payment_method, date_from, date_to]):
            writer.writerow([f"# 필터 조건 적용됨 - 생성일시: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"])
            if category:
                writer.writerow([f"# 카테고리: {category}"])
            if payment_method:
                writer.writerow([f"# 결제수단: {payment_method}"])
            if date_from:
                writer.writerow([f"# 시작일: {date_from}"])
            if date_to:
                writer.writerow([f"# 종료일: {date_to}"])
            writer.writerow([])
        
        # Write headers
        writer.writerow([
            "날짜", "금액(원)", "금액(엔)", "카테고리", "설명", "결제수단", "등록시간"
        ])
        
        # Write data
        for expense in expenses:
            jpy_amount = round(expense.amount / exchange_rate) if exchange_rate > 0 else 0
            writer.writerow([
                expense.date,
                f"{expense.amount:,.0f}",
                f"¥{jpy_amount:,}",
                expense.category,
                expense.description,
                expense.payment_method,
                expense.timestamp.strftime("%Y-%m-%d %H:%M:%S") if expense.timestamp else ""
            ])
        
        # Add summary at the end
        total_amount = sum(expense.amount for expense in expenses)
        total_jpy = round(total_amount / exchange_rate) if exchange_rate > 0 else 0
        writer.writerow([])
        writer.writerow([f"총 {len(expenses)}건", f"{total_amount:,.0f}원", f"¥{total_jpy:,}", "", "", "", ""])
        writer.writerow([f"환율 정보: 1엔 = {exchange_rate:.2f}원", "", "", "", "", "", ""])
        
        # Prepare response
        csv_content = output.getvalue()
        output.close()
        
        filename = f"japan_expenses_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
        
        from fastapi.responses import Response
        return Response(
            content=csv_content.encode('utf-8-sig'),  # BOM for Excel compatibility
            media_type="text/csv; charset=utf-8",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"CSV export failed: {str(e)}")

@app.get("/api/export/excel")
async def export_expenses_excel(
    category: Optional[str] = None,
    payment_method: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    db: Session = Depends(get_db)
):
    """Export expenses data as Excel file."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
        from datetime import datetime
        import io
        
        # Get filtered expenses
        expenses = ExpenseService.get_filtered_expenses(
            db=db,
            category=category,
            payment_method=payment_method,
            date_from=date_from,
            date_to=date_to,
            sort_by="date",
            sort_order="desc"
        )
        
        # Get current exchange rate
        try:
            exchange_rate = exchange_service.get_jpy_to_krw_rate()
        except:
            exchange_rate = 9.5  # Fallback rate
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "일본여행 지출내역"
        
        row = 1
        
        # Add filter information
        if any([category, payment_method, date_from, date_to]):
            ws[f'A{row}'] = f"필터 조건 - 생성일시: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
            ws[f'A{row}'].font = Font(bold=True)
            row += 1
            
            if category:
                ws[f'A{row}'] = f"카테고리: {category}"
                row += 1
            if payment_method:
                ws[f'A{row}'] = f"결제수단: {payment_method}"
                row += 1
            if date_from:
                ws[f'A{row}'] = f"시작일: {date_from}"
                row += 1
            if date_to:
                ws[f'A{row}'] = f"종료일: {date_to}"
                row += 1
            row += 1
        
        # Headers
        headers = ["날짜", "금액(원)", "금액(엔)", "카테고리", "설명", "결제수단", "등록시간"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="E6E6FA", end_color="E6E6FA", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        row += 1
        
        # Data rows
        for expense in expenses:
            jpy_amount = round(expense.amount / exchange_rate) if exchange_rate > 0 else 0
            
            ws.cell(row=row, column=1, value=expense.date)
            ws.cell(row=row, column=2, value=expense.amount)
            ws.cell(row=row, column=3, value=jpy_amount)
            ws.cell(row=row, column=4, value=expense.category)
            ws.cell(row=row, column=5, value=expense.description)
            ws.cell(row=row, column=6, value=expense.payment_method)
            ws.cell(row=row, column=7, value=expense.timestamp.strftime("%Y-%m-%d %H:%M:%S") if expense.timestamp else "")
            row += 1
        
        # Summary
        total_amount = sum(expense.amount for expense in expenses)
        total_jpy = round(total_amount / exchange_rate) if exchange_rate > 0 else 0
        
        row += 1
        ws.cell(row=row, column=1, value=f"총 {len(expenses)}건")
        ws.cell(row=row, column=2, value=total_amount)
        ws.cell(row=row, column=3, value=total_jpy)
        
        # Make summary row bold
        for col in range(1, 4):
            ws.cell(row=row, column=col).font = Font(bold=True)
        
        row += 1
        ws.cell(row=row, column=1, value=f"환율 정보: 1엔 = {exchange_rate:.2f}원")
        ws.cell(row=row, column=1).font = Font(italic=True)
        
        # Auto adjust column widths
        for column_cells in ws.columns:
            length = max(len(str(cell.value or "")) for cell in column_cells)
            ws.column_dimensions[column_cells[0].column_letter].width = min(max(length + 2, 10), 50)
        
        # Save to bytes
        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        filename = f"japan_expenses_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        from fastapi.responses import Response
        return Response(
            content=excel_buffer.getvalue(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Excel export failed: {str(e)}")

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)